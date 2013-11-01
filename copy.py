#coding = utf-8
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
data_col_start_init = raw_input('from which column of data you want to copy:')
data_col_end_init = raw_input('to which column of data you want to copy:')
data_row_start_init = raw_input('from which row of data you want to copy:')
data_row_end_init =  raw_input('to which row of data you want to copy:')
col_insert_init = raw_input('enter the Column of Position you want to insert to:')
raw_insert_init = raw_input('enter the Row of Position you want to insert to:')
wb_from = load_workbook('day3.xlsx')   #the source excel
ws_from = wb_from.get_active_sheet()
wb_to = Workbook()
ws_to = wb_to.worksheets[0]
ws_to.title = 'first'
col_insert = int(col_insert_init) - 1          

#copy
for coll in xrange(int(data_col_start_init) - 1,int(data_col_end_init)):  #column range of data you want to copy
    row_insert = int(raw_insert_init) - 1       
    for roww in xrange(int(data_row_start_init) - 1,int(data_row_end_init)):   #row range of data you want to copy
        ws_to.cell(row=row_insert, column=col_insert).value = ws_from.cell(row=roww, column=coll).value
        row_insert +=1
    col_insert += 1

wb_to.save('day3copy.xlsx')   #the destination excel
